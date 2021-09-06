# -*- coding: utf8 -*-
from selenium import webdriver
from openpyxl.styles import Alignment, Font, PatternFill
from datetime import datetime
import openpyxl

#Функция для получения текущей даты и времени
def log_data_time():
    current_time = datetime.now()
    time_log = current_time.strftime("%d.%m.%Y %H:%M:%S")
    return time_log

#Логируем действие в файл
with open('info.log', 'a', encoding='utf-8') as f:
    print('-----------------------------------------', file=f)
    print(f"{log_data_time()} Скрипт запущен\n", file=f)

print("Скрипт запущен...")

#Загружаем данные из файла 'Справочник контрагентов.xlsx'
wb = openpyxl.load_workbook('Справочник контрагентов.xlsx')
sheet_ranges = wb['TDSheet']

# #Создаем файл Excel
wb = openpyxl.Workbook()
ws = wb.active

#Называем лист Excel
ws.title = 'Список'

#Задаем заголовки
ws['A1'] = 'Наименование '
ws['B1'] = 'ИНН'
ws['C1'] = 'ЮЛ'
ws['D1'] = 'НДС'
ws['E1'] = 'Является ли должником'
ws['F1'] = 'Является ли банкротом'
ws['G1'] = 'Вид деятельности:'
ws['H1'] = 'Адрес:'
ws['I1'] = 'Дата регистрации:'
ws['J1'] = 'Номер регистрации:'
ws['K1'] = 'ОКПО:'
ws['L1'] = 'СОАТО:'
ws['M1'] = 'ОПФ:'
ws['N1'] = 'ОКЭД:'
ws['O1'] = 'СООГУ:'

#Задаем ширину столбцов
ws.column_dimensions["A"].width = 80 #Наименование
ws.column_dimensions["B"].width = 12 #ИНН
ws.column_dimensions["C"].width = 30 #ЮЛ
ws.column_dimensions["D"].width = 35 #НДС
ws.column_dimensions["E"].width = 35 #Является ли должником
ws.column_dimensions["F"].width = 25 #Является ли банкротом
ws.column_dimensions["G"].width = 30 #Вид деятельности
ws.column_dimensions["H"].width = 45 #Адрес
ws.column_dimensions["I"].width = 25 #Дата регистрации
ws.column_dimensions["J"].width = 20 #Номер регистрации
ws.column_dimensions["K"].width = 50 #ОКПО
ws.column_dimensions["L"].width = 30 #СОАТО
ws.column_dimensions["M"].width = 50 #ОПФ
ws.column_dimensions["N"].width = 90 #ОКЭД
ws.column_dimensions["O"].width = 130 #СООГУ

#Выравниваем заголовок по центру
ws['A1'].alignment = Alignment(horizontal="center")
ws['B1'].alignment = Alignment(horizontal="center")
ws['C1'].alignment = Alignment(horizontal="center")
ws['D1'].alignment = Alignment(horizontal="center")
ws['E1'].alignment = Alignment(horizontal="center")
ws['F1'].alignment = Alignment(horizontal="center")
ws['G1'].alignment = Alignment(horizontal="center")
ws['H1'].alignment = Alignment(horizontal="center")
ws['I1'].alignment = Alignment(horizontal="center")
ws['J1'].alignment = Alignment(horizontal="center")
ws['K1'].alignment = Alignment(horizontal="center")
ws['L1'].alignment = Alignment(horizontal="center")
ws['M1'].alignment = Alignment(horizontal="center")
ws['N1'].alignment = Alignment(horizontal="center")
ws['O1'].alignment = Alignment(horizontal="center")

#Выделяем заголовок "жирным текстом"
ws['A1'].font = Font(bold=True)
ws['B1'].font = Font(bold=True)
ws['C1'].font = Font(bold=True)
ws['D1'].font = Font(bold=True)
ws['E1'].font = Font(bold=True)
ws['F1'].font = Font(bold=True)
ws['G1'].font = Font(bold=True)
ws['H1'].font = Font(bold=True)
ws['I1'].font = Font(bold=True)
ws['J1'].font = Font(bold=True)
ws['K1'].font = Font(bold=True)
ws['L1'].font = Font(bold=True)
ws['M1'].font = Font(bold=True)
ws['N1'].font = Font(bold=True)
ws['O1'].font = Font(bold=True)

#Фон для заголовка таблицы
ws['A1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['B1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['C1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['D1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['E1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['F1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['G1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['H1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['I1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['J1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['K1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['L1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['M1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['N1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")
ws['O1'].fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type = "solid")

#Заморозка первой строки, столбца
wb['Список'].freeze_panes = "B2"

#Добавляем автофильтр
ws.auto_filter.ref = ws.dimensions

#Счётчик ячеек
count = 2

#Cчётчик ошибок
count_e = 1

for row in sheet_ranges.iter_rows():
    if row[0].value != None and row[0].value != 'ИНН':

        try:
            search = 'https://my2.soliq.uz/main/info/personal?searchtin=' + str(row[0].value).replace('.', '').replace(' ','')
            options = webdriver.ChromeOptions()
            options.add_argument('--headless')
            driver = webdriver.Chrome(executable_path=r'chromedriver.exe', options=options)
            driver.get(search)

            inn = str(driver.find_element_by_xpath('//*[@id="tinHolder"]').text).replace('ИНН: ', '')
            org = driver.find_element_by_xpath('//*[@id="nameinfo"]/h4').text
            entity = driver.find_element_by_xpath('//*[@id="nameinfo"]/p').text
            driver.implicitly_wait(1)
            tax = driver.find_element_by_xpath('//*[@id="ndsStatus"]').text
            debtor = driver.find_element_by_xpath('//*[@id="debtorStatus"]').text
            bankrupt = driver.find_element_by_xpath('//*[@id="bankrotStatus"]').text
            activity = driver.find_element_by_xpath('//*[@id="infotable"]/tr[1]/td').text
            address = driver.find_element_by_xpath('//*[@id="infotable"]/tr[2]/td').text
            reg_date = driver.find_element_by_xpath('//*[@id="infotable"]/tr[3]/td').text
            reg_number = driver.find_element_by_xpath('//*[@id="infotable"]/tr[4]/td').text
            okpo = driver.find_element_by_xpath('//*[@id="infotable"]/tr[5]/td').text
            soato = driver.find_element_by_xpath('//*[@id="infotable"]/tr[6]/td').text
            opf = driver.find_element_by_xpath('//*[@id="infotable"]/tr[7]/td').text
            oked = driver.find_element_by_xpath('//*[@id="infotable"]/tr[8]/td').text
            soogu = driver.find_element_by_xpath('//*[@id="infotable"]/tr[9]/td').text

            generated_html = driver.page_source
            driver.quit()

            ws['A' + str(count)] = org
            ws['B' + str(count)] = inn
            ws['C' + str(count)] = entity
            ws['D' + str(count)] = tax
            ws['E' + str(count)] = debtor
            ws['F' + str(count)] = bankrupt
            ws['G' + str(count)] = activity
            ws['H' + str(count)] = address
            ws['I' + str(count)] = reg_date
            ws['J' + str(count)] = reg_number
            ws['K' + str(count)] = okpo
            ws['L' + str(count)] = soato
            ws['M' + str(count)] = opf
            ws['N' + str(count)] = oked
            ws['O' + str(count)] = soogu
            ws['A' + str(count)].font = Font(size = '12', bold=True)
            count += 1

        except:
            with open('info.log', 'a', encoding='utf-8') as f:
                print(f'{count_e}.ИНН не найден: {row[0].value} {row[3].value}', file=f)
                count_e += 1

wb.save('Результат проверки.xlsx')

with open('info.log', 'a', encoding='utf-8') as f:
    print(f"\n{log_data_time()} Скрипт отработал успешно!!!", file=f)

print('Скрипт выполнен')